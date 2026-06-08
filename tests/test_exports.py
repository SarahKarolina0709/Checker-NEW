# -*- coding: utf-8 -*-
"""Tests fuer nicegui_app.exports."""
import os
import sys
import tempfile
import shutil
import zipfile
from dataclasses import dataclass

import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from nicegui_app import exports as E


@dataclass
class _F:
    severity: str = 'major'
    code: str = 'X1'
    message: str = 'msg'
    source_text: str = ''
    target_text: str = ''
    category: str = ''


def _findings():
    return [
        _F('critical', 'C1', 'Kritischer Fehler', 'src1', 'tgt1', 'numbers'),
        _F('major', 'M1', 'Wichtig <html> & co', 'src2', 'tgt2', 'terminology'),
        _F('minor', 'm1', 'Klein', '', '', 'style'),
    ]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
class TestHelpers:
    def test_xml_clean_strips_control_chars(self):
        assert E._xml_clean('a\x00b\x07c') == 'abc'

    def test_xml_clean_keeps_tabs_newlines(self):
        assert E._xml_clean('a\tb\nc\rd') == 'a\tb\nc\rd'

    def test_xml_clean_passthrough_non_str(self):
        assert E._xml_clean(42) == 42
        assert E._xml_clean(None) is None

    def test_xml_escape(self):
        assert E._xml_escape('<a & b>') == '&lt;a &amp; b&gt;'

    def test_xml_escape_empty(self):
        assert E._xml_escape('') == ''
        assert E._xml_escape(None) == ''

    def test_truncate_short_unchanged(self):
        assert E._truncate('kurz') == 'kurz'

    def test_truncate_empty_stays_empty(self):
        # Aufrufer prueft `if text:` -> leeres Feld wird nicht gerendert
        assert E._truncate('') == ''
        assert E._truncate(None) == ''

    def test_truncate_long_marks_cut(self):
        out = E._truncate('A' * 1000, limit=600)
        assert out.endswith('(gekürzt)')
        assert len(out) < 1000
        assert out.startswith('A' * 600)


# ---------------------------------------------------------------------------
# TXT
# ---------------------------------------------------------------------------
class TestTxt:
    def setup_method(self):
        self.tmp = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_basic(self):
        path = E.export_txt(_findings(), 75, self.tmp)
        assert os.path.isfile(path)
        content = open(path, encoding='utf-8').read()
        assert 'Score: 75/100' in content
        assert 'Kritischer Fehler' in content
        assert 'src1' in content
        assert 'tgt1' in content

    def test_creates_output_dir(self):
        new_dir = os.path.join(self.tmp, 'new', 'sub')
        path = E.export_txt(_findings(), 50, new_dir)
        assert os.path.isfile(path)

    def test_empty_findings(self):
        path = E.export_txt([], 100, self.tmp)
        assert os.path.isfile(path)
        assert 'Befunde: 0' in open(path, encoding='utf-8').read()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
class TestExcel:
    def setup_method(self):
        self.tmp = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_basic(self):
        try:
            import openpyxl  # noqa
        except ImportError:
            pytest.skip('openpyxl missing')
        path = E.export_excel(_findings(), self.tmp)
        assert os.path.isfile(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Header + 3 Datenzeilen
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4
        assert rows[0][0] == 'Nr'
        assert rows[1][2] == 'C1'

    def test_strips_illegal_xml_chars(self):
        try:
            import openpyxl  # noqa
        except ImportError:
            pytest.skip('openpyxl missing')
        bad = [_F('major', 'X', 'hi\x00there', 'a\x07b', '', '')]
        path = E.export_excel(bad, self.tmp)
        wb = openpyxl.load_workbook(path)
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[1][4] == 'hithere'  # message
        assert rows[1][5] == -1         # segment_index (default)
        assert rows[1][6] == 'ab'       # source


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
class TestPdf:
    def setup_method(self):
        self.tmp = tempfile.mkdtemp()

    def teardown_method(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_basic(self):
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        path = E.export_pdf(_findings(), 80, self.tmp)
        assert os.path.isfile(path)
        # PDF Header-Magic
        with open(path, 'rb') as f:
            assert f.read(4) == b'%PDF'

    def test_special_chars_no_crash(self):
        # Bug-Regression: <, &, > in messages duerfen ReportLab nicht crashen
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        path = E.export_pdf([_F('major', 'X<>&', 'a < b & c > d')], 50, self.tmp)
        assert os.path.isfile(path)

    def test_empty_findings_no_crash(self):
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        path = E.export_pdf([], 100, self.tmp)
        assert os.path.isfile(path)
        with open(path, 'rb') as f:
            assert f.read(4) == b'%PDF'

    def test_all_severities_render(self):
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        findings = [
            _F('critical', 'C', 'krit', 'Quelle A', 'Ziel A'),
            _F('major', 'M', 'wichtig', 'Quelle B', 'Ziel B'),
            _F('minor', 'm', 'klein', 'Quelle C', 'Ziel C'),
            _F('info', 'i', 'hinweis', 'Quelle D', 'Ziel D'),
        ]
        path = E.export_pdf(findings, 30, self.tmp)
        assert os.path.isfile(path)

    def test_bilingual_content_grows_file(self):
        """PDF mit Quell-/Zieltext ist groesser als ohne (Inhalt landet im PDF)."""
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        # Getrennte Verzeichnisse: _timestamp() hat nur Sekunden-Aufloesung,
        # gleiche Sekunde -> gleicher Dateiname -> Kollision.
        d1 = os.path.join(self.tmp, 'a')
        d2 = os.path.join(self.tmp, 'b')
        minimal = E.export_pdf([_F('major', 'X', 'kurz')], 80, d1)
        with_text = E.export_pdf(
            [_F('major', 'X', 'kurz', 'A' * 400, 'B' * 400)], 80, d2)
        assert os.path.getsize(with_text) > os.path.getsize(minimal)

    def test_multipage_build_no_crash(self):
        """Viele lange Befunde -> mehrseitiger Bericht (uebt die Seiten-Fusszeile)."""
        try:
            import reportlab  # noqa
        except ImportError:
            pytest.skip('reportlab missing')
        findings = [
            _F('major', f'CODE_{i}', f'Befund Nummer {i} ' * 5,
               'Q' * 700, 'Z' * 700)
            for i in range(60)
        ]
        path = E.export_pdf(findings, 25, self.tmp)
        assert os.path.isfile(path)
        with open(path, 'rb') as f:
            assert f.read(4) == b'%PDF'


# ---------------------------------------------------------------------------
# Korrekturpaket
# ---------------------------------------------------------------------------
class TestZipPackage:
    def setup_method(self):
        self.tmp = tempfile.mkdtemp()
        # Ein paar Demo-Files
        self.src1 = os.path.join(self.tmp, 'src', 'a.txt')
        self.tgt1 = os.path.join(self.tmp, 'tgt', 'a.txt')
        os.makedirs(os.path.dirname(self.src1))
        os.makedirs(os.path.dirname(self.tgt1))
        open(self.src1, 'w', encoding='utf-8').write('S')
        open(self.tgt1, 'w', encoding='utf-8').write('T')

    def teardown_method(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_creates_zip(self):
        out = E.export_correction_package(
            _findings(), 70, [self.src1], [self.tgt1], self.tmp,
        )
        assert out and os.path.isfile(out)
        with zipfile.ZipFile(out) as zf:
            names = zf.namelist()
            assert any(n.startswith('01_Ausgangstexte/') for n in names)
            assert any(n.startswith('02_Uebersetzungen/') for n in names)
            assert 'ANSCHREIBEN.txt' in names
            assert 'KORREKTUREN.txt' in names
            assert any(n.startswith('bericht/') for n in names)

    def test_filename_collision_dedup(self):
        # Zweite source-Datei mit gleichem Basename in anderem Ordner
        other = os.path.join(self.tmp, 'src2', 'a.txt')
        os.makedirs(os.path.dirname(other))
        open(other, 'w', encoding='utf-8').write('S2')
        out = E.export_correction_package(
            _findings(), 70, [self.src1, other], [], self.tmp,
        )
        with zipfile.ZipFile(out) as zf:
            ausgangs = [n for n in zf.namelist() if n.startswith('01_Ausgangstexte/')]
            assert len(ausgangs) == 2
            # Zweiter darf NICHT gleichen Pfad haben
            assert len(set(ausgangs)) == 2

    def test_skips_missing_files(self):
        out = E.export_correction_package(
            _findings(), 70, [self.src1, '/nonexistent/x.txt'], [], self.tmp,
        )
        with zipfile.ZipFile(out) as zf:
            ausgangs = [n for n in zf.namelist() if n.startswith('01_Ausgangstexte/')]
            assert len(ausgangs) == 1

    def test_empty_files_lists(self):
        out = E.export_correction_package(_findings(), 50, [], [], self.tmp)
        assert out and os.path.isfile(out)
        with zipfile.ZipFile(out) as zf:
            assert 'KORREKTUREN.txt' in zf.namelist()
            assert 'ANSCHREIBEN.txt' in zf.namelist()


class TestCorrectionGrouping:
    """Prioritaets-Gruppierung der Korrekturliste (Namen/Zahlen/Vollst./Rest)."""

    def test_priority_group_mapping(self):
        assert E._priority_group(_F(code='NUMBER_MISSING')) == E._GRP_ZAHLEN
        assert E._priority_group(_F(code='COMPANY_SUFFIX_TRANSLATED')) == E._GRP_NAMEN
        assert E._priority_group(_F(code='GLOSSARY_VIOLATION')) == E._GRP_NAMEN
        assert E._priority_group(
            _F(code='COMPLETENESS_TOO_SHORT', category='completeness')) == E._GRP_VOLLST
        assert E._priority_group(
            _F(code='UNTRANSLATED_SEGMENT', category='completeness')) == E._GRP_VOLLST
        assert E._priority_group(_F(code='STYLE_PASSIVE')) == E._GRP_SONST

    def test_groups_ordered_by_priority(self):
        findings = [
            _F('minor', 'STYLE_X', 'stil', category='style'),
            _F('major', 'COMPLETENESS_TOO_SHORT', 'kurz', category='completeness'),
            _F('major', 'NUMBER_MISSING', 'zahl'),
            _F('critical', 'COMPANY_NAME_MISSING', 'name'),
        ]
        txt = E._build_corrections(findings, 50)
        i_namen = txt.index('NAMEN & EIGENNAMEN')
        i_zahlen = txt.index('ZAHLEN')
        i_vollst = txt.index('VOLLSTÄNDIGKEIT')
        i_sonst = txt.index('WEITERE PUNKTE')
        assert i_namen < i_zahlen < i_vollst < i_sonst

    def test_cover_note_lists_priorities(self):
        note = E._build_cover_note(_findings(), 70)
        assert 'NAMEN' in note and 'ZAHLEN' in note and 'VOLLSTÄNDIGKEIT' in note
        assert 'KORREKTUREN.txt' in note

    def test_severity_summary_singular_plural(self):
        assert E._severity_summary([_F('minor')]).endswith('1 Hinweis')
        assert E._severity_summary([_F('minor'), _F('info')]).endswith('2 Hinweise')
