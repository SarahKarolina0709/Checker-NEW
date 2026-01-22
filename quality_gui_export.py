"""Quality GUI Export-Modul (DEPRECATED)

DEPRECATED: Dieses Modul wurde durch `quality_gui_reporting.py` ersetzt.
Beibehalten nur für Rückwärtskompatibilität älterer Skripte.
Neue Implementierungen sollen ausschließlich das Reporting-Modul verwenden.

Historische Funktionen:
 - export_analysis / export_multi
 - export_txt / export_pdf / export_excel

Migrationshinweis:
 - Verwende `QualityGuiReporting.perform_export(app, format)` für neue Exporte.
 - Events & Result-Objekte werden nun zentral dort behandelt.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, Sequence, Callable, Iterable, Optional
from dataclasses import dataclass

logger = logging.getLogger(__name__)

SUPPORTED_FORMATS = ("txt", "pdf", "xlsx")


@dataclass
class ExportResult:
    path: Path
    format: str
    success: bool = True
    fallback: bool = False
    message: str | None = None

EventCallback = Callable[[str, Dict[str, Any]], None]


def export_analysis(
    report: Dict[str, Any] | Sequence[Dict[str, Any]],
    path: str | Path,
    format: str = "txt",
    *,
    event_cb: Optional[EventCallback] = None,
    include_charts: bool = True,
    force_suffix: bool = True,
) -> Path:
    """High-Level Export mit optionalem Event Callback.

    Parameters
    ----------
    report : Mapping | Sequence[Mapping]
        Analyse-Daten.
    path : str | Path
        Basis-Pfad oder Zieldatei.
    format : str
        Zielformat ('txt','pdf','xlsx').
    event_cb : Optional[Callable]
        Callback(event_name, context_dict) – context enthält u.a. format, target, success.
    include_charts : bool
        Bei PDF optionalen Metrik-Chart anhängen.
    force_suffix : bool
        True (Default) erzwingt Dateiendung passend zum Format. False lässt benutzerdefinierten
        Suffix bestehen (nur wenn bereits vorhanden); fehlt Suffix komplett wird trotzdem
        richtig ergänzt.

    Events
    ------
    export.started  -> {format, target}
    export.done     -> {format, target, success, fallback, message?}
    """
    # Früh prüfen, damit export_multi() nicht leer zurückkommt
    fmt_l = format.lower().strip()
    if fmt_l not in SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported format '{format}'. Supported: {', '.join(SUPPORTED_FORMATS)}")

    res = export_multi(
        report, path, [format], event_cb=event_cb, include_charts=include_charts, force_suffix=force_suffix
    )
    return res[0].path


def export_multi(
    report: Dict[str, Any] | Sequence[Dict[str, Any]],
    base_path: str | Path,
    formats: Iterable[str],
    *,
    event_cb: Optional[EventCallback] = None,
    include_charts: bool = True,
    force_suffix: bool = True,
) -> list[ExportResult]:
    """Export mehrerer Formate.

    force_suffix steuert, ob vorhandene vom Benutzer gesetzte Suffixe überschrieben werden.
    """
    results: list[ExportResult] = []
    for fmt in formats:
        fmt_l = fmt.lower().strip()
        if fmt_l not in SUPPORTED_FORMATS:
            logger.warning("Unsupported format '%s' – skip", fmt_l)
            continue
        target = Path(base_path)
        # Suffix Handling
        if force_suffix:
            if target.suffix.lower().lstrip('.') != fmt_l:
                # Wenn kein Suffix -> einfach anhängen. Wenn anders -> ersetzen.
                if target.suffix:
                    target = target.with_suffix(f'.{fmt_l}')
                else:
                    target = target.with_name(target.name + f'.{fmt_l}')
        else:
            # force_suffix False: nur ergänzen falls kein Suffix existiert
            if not target.suffix:
                target = target.with_name(target.name + f'.{fmt_l}')
        if event_cb:
            _safe_event(event_cb, 'export.started', format=fmt_l, target=str(target))
        try:
            # Zielordner sicherstellen
            try:
                target.parent.mkdir(parents=True, exist_ok=True)
            except Exception:
                logger.debug("Konnte Zielverzeichnis nicht anlegen: %s", target, exc_info=True)
            fallback = False
            success = True
            if fmt_l == 'txt':
                export_txt(report, target)
            elif fmt_l == 'pdf':
                ok = export_pdf(report, target, include_charts=include_charts)
                if not ok:
                    fallback = True
                    fb = target.with_suffix('.txt')
                    export_txt(report, fb)
                    target = fb
            elif fmt_l == 'xlsx':
                ok = export_excel(report, target)
                if not ok:
                    fallback = True
                    fb = target.with_suffix('.txt')
                    export_txt(report, fb)
                    target = fb
            msg = None
            if fallback:
                msg = f"{fmt_l.upper()} nicht verfügbar, Fallback zu TXT"
            logger.info("Export erfolgreich: %s (fallback=%s)", target, fallback)
            results.append(ExportResult(path=target, format=fmt_l, success=success, fallback=fallback, message=msg))
            if event_cb:
                _safe_event(event_cb, 'export.done', format=fmt_l, target=str(target), success=success, fallback=fallback, message=msg)
        except Exception as e:  # pragma: no cover
            logger.exception("Export fehlgeschlagen (%s)", fmt_l)
            results.append(ExportResult(path=target, format=fmt_l, success=False, fallback=False, message=str(e)))
            if event_cb:
                _safe_event(event_cb, 'export.done', format=fmt_l, target=str(target), success=False, error=str(e))
    return results
def _safe_event(cb: EventCallback, name: str, **ctx: Any) -> None:
    try:
        cb(name, ctx)
    except Exception:  # pragma: no cover
        logger.debug("Event Callback Fehler ignoriert", exc_info=True)



def _normalize(report: Dict[str, Any] | Sequence[Dict[str, Any]]) -> list[Dict[str, Any]]:
    if isinstance(report, dict):
        return [report]
    return list(report)


def export_txt(report: Dict[str, Any] | Sequence[Dict[str, Any]], target: Path) -> None:
    rows = _normalize(report)
    lines = ["QUALITY ANALYSIS EXPORT", "=" * 28, ""]
    # Konsistente Reihenfolge: Keys aus erstem Row, weitere anhängen
    first_keys = list(rows[0].keys()) if rows else []
    _seen = set(first_keys)
    for r in rows[1:]:
        for k in r.keys():
            if k not in _seen:
                first_keys.append(k); _seen.add(k)
    # Versuche gruppierte Ausgabe: Aggregate / Pairs / Plugins
    try:
        aggregates = [r for r in rows if isinstance(r, dict) and r.get('row_type') == 'aggregate']
        pairs = [r for r in rows if isinstance(r, dict) and r.get('row_type') == 'pair']
        plugins = [r for r in rows if isinstance(r, dict) and r.get('row_type') == 'plugin']
        if aggregates or pairs or plugins:
            if aggregates:
                lines.append("AGGREGATE METRICS")
                lines.append("-" * 18)
                for ag in aggregates:
                    for k, v in sorted(ag.items()):
                        if k == 'row_type':
                            continue
                        lines.append(f"  {k}: {v}")
                lines.append("")
            if pairs:
                lines.append("FILE PAIRS")
                lines.append("-" * 10)
                for p in pairs:
                    lines.append(f"Pair {p.get('id','?')}")
                    for k, v in sorted(p.items()):
                        if k in ('row_type','id'): continue
                        lines.append(f"  {k}: {v}")
                    lines.append("")
            if plugins:
                lines.append("PLUGIN RESULTS")
                lines.append("-" * 14)
                for pr in plugins:
                    lines.append(f"Rule {pr.get('rule')}: passed={pr.get('passed')}")
                    for k, v in sorted(pr.items()):
                        if k in ('row_type','rule','passed'): continue
                        lines.append(f"  {k}: {v}")
                    lines.append("")
        else:  # Fallback Standard Format (mit stabiler Key-Reihenfolge)
            for idx, row in enumerate(rows, 1):
                lines.append(f"Item {idx}:")
                for k in first_keys:
                    v = row.get(k, "")
                    lines.append(f"  - {k}: {v}")
                lines.append("")
    except Exception:
        for idx, row in enumerate(rows, 1):  # final fallback
            lines.append(f"Item {idx}:")
            for k, v in sorted(getattr(row, 'items', lambda: [])()):
                lines.append(f"  - {k}: {v}")
            lines.append("")
    try:
        target.write_text("\n".join(lines), encoding='utf-8')
    except Exception as e:  # pragma: no cover
        logger.error("TXT Export fehlgeschlagen: %s", e)
        raise


# Rückwärtskompatibel
_export_txt = export_txt


def export_pdf(report: Dict[str, Any] | Sequence[Dict[str, Any]], target: Path, *, include_charts: bool = True) -> bool:
    """PDF Export mit professioneller Darstellung von Befunden.

    Rückgabe False wenn reportlab fehlt.
    """
    try:  # Basis-Import
        from reportlab.lib.pagesizes import A4  # type: ignore
        from reportlab.pdfgen import canvas  # type: ignore
        from reportlab.lib.utils import ImageReader  # type: ignore
        from reportlab.lib import colors  # type: ignore
    except Exception:
        logger.warning("reportlab nicht verfügbar – PDF Export übersprungen")
        return False

    # Daten extrahieren
    rows = _normalize(report)
    data = rows[0] if rows else {}
    
    # Findings und Metriken extrahieren
    findings = []
    metrics = {}
    summary = {}
    score = None
    
    if isinstance(data, dict):
        # Findings aus verschiedenen Quellen sammeln
        if isinstance(data.get('findings'), list):
            findings = data.get('findings', [])
        # Auch issues_phase1, phase2, phase3 sammeln
        for phase_key in ['issues_phase1', 'issues_phase2', 'issues_phase3']:
            phase_issues = data.get(phase_key)
            if isinstance(phase_issues, list):
                findings.extend(phase_issues)
        
        metrics = data.get('metrics', {}) if isinstance(data.get('metrics'), dict) else {}
        summary = data.get('summary', {}) if isinstance(data.get('summary'), dict) else {}
        
        # Score ermitteln
        score = summary.get('quality_score') or summary.get('overall_score_norm') or metrics.get('overall_score')
        if isinstance(score, (int, float)):
            score = score * 100 if score <= 1 else score

    chart_path: Optional[Path] = None
    if include_charts:
        chart_path = _maybe_build_metrics_chart(report)

    # Professionelles PDF mit platypus
    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
        from reportlab.lib.units import cm  # type: ignore
        from xml.sax.saxutils import escape as _rl_escape
        import datetime
        styles = getSampleStyleSheet()
        
        # Custom Styles - nur hinzufügen wenn nicht vorhanden
        def _add_style_safe(name, **kwargs):
            if name not in styles.byName:
                styles.add(ParagraphStyle(name=name, **kwargs))
        
        _add_style_safe('Title_DE', parent=styles['Heading1'], fontSize=18, spaceAfter=12, textColor=colors.HexColor('#1F4E79'))
        _add_style_safe('Section', parent=styles['Heading2'], fontSize=14, spaceBefore=16, spaceAfter=8, textColor=colors.HexColor('#374151'))
        _add_style_safe('FindingTitle', parent=styles['Normal'], fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#1F4E79'))
        _add_style_safe('FindingText', parent=styles['Normal'], fontSize=10, leftIndent=12, textColor=colors.HexColor('#374151'))
        _add_style_safe('ErrorText', parent=styles['Normal'], fontSize=10, leftIndent=12, textColor=colors.HexColor('#DC2626'))
        _add_style_safe('SourceTarget', parent=styles['Normal'], fontSize=9, leftIndent=20, backColor=colors.HexColor('#F3F4F6'), borderPadding=4)
        
        story = []
        
        # Titel
        story.append(Paragraph("Qualitätsanalyse-Bericht", styles['Title_DE']))
        story.append(Paragraph(f"Erstellt am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Score-Bereich
        if score is not None:
            score_color = '#10B981' if score >= 75 else '#F59E0B' if score >= 50 else '#DC2626'
            story.append(Paragraph("Gesamtbewertung", styles['Section']))
            story.append(Paragraph(f"<font color='{score_color}' size='24'><b>{int(round(score))}/100</b></font>", styles['Normal']))
            
            # Bewertungstext
            if score >= 90:
                quality_text = "Ausgezeichnete Qualität"
            elif score >= 75:
                quality_text = "Gute Qualität"
            elif score >= 50:
                quality_text = "Verbesserungsbedarf"
            else:
                quality_text = "Erhebliche Mängel"
            story.append(Paragraph(f"<i>{quality_text}</i>", styles['Normal']))
            story.append(Spacer(1, 16))
        
        # Chart einfügen
        if chart_path and chart_path.exists():
            try:
                story.append(Image(str(chart_path), width=280, height=180))
                story.append(Spacer(1, 12))
            except Exception:
                pass
        
        # Metriken-Übersicht
        if metrics:
            story.append(Paragraph("Kennzahlen", styles['Section']))
            metric_labels = {
                'numeric_consistency': 'Numerische Konsistenz',
                'untranslated_ratio': 'Unübersetzter Anteil',
                'avg_length_ratio': 'Durchschn. Längenverhältnis',
                'segments_total': 'Segmente gesamt',
                'segments_translated': 'Übersetzte Segmente'
            }
            for key, label in metric_labels.items():
                val = metrics.get(key)
                if val is not None:
                    if isinstance(val, float) and val <= 1:
                        val_str = f"{val*100:.1f}%"
                    else:
                        val_str = str(val)
                    story.append(Paragraph(f"<b>{label}:</b> {val_str}", styles['Normal']))
            story.append(Spacer(1, 16))
        
        # Befunde (Findings)
        if findings:
            story.append(Paragraph(f"Befunde ({len(findings)} gefunden)", styles['Section']))
            
            # Regel-Namen für bessere Lesbarkeit
            rule_friendly_names = {
                'terminology_mismatch': 'Terminologie-Abweichung',
                'missing_translation': 'Fehlende Übersetzung',
                'untranslated_segment': 'Unübersetztes Segment',
                'number_mismatch': 'Zahlen stimmen nicht überein',
                'length_anomaly': 'Ungewöhnliche Längenabweichung',
                'grammar_error': 'Grammatikfehler',
                'spelling_error': 'Rechtschreibfehler',
                'consistency_error': 'Konsistenzfehler',
                'punctuation_mismatch': 'Zeichensetzungsfehler',
                'capitalization_error': 'Groß-/Kleinschreibung'
            }
            
            severity_colors = {
                'critical': '#DC2626',
                'major': '#F59E0B',
                'minor': '#3B82F6',
                'info': '#6B7280'
            }
            severity_labels = {
                'critical': 'Kritisch',
                'major': 'Wesentlich',
                'minor': 'Geringfügig',
                'info': 'Hinweis'
            }
            
            # Gruppiere nach Schweregrad
            grouped = {'critical': [], 'major': [], 'minor': [], 'info': []}
            for f in findings:
                sev = f.get('severity', 'info') if isinstance(f, dict) else 'info'
                if sev not in grouped:
                    sev = 'info'
                grouped[sev].append(f)
            
            for severity in ['critical', 'major', 'minor', 'info']:
                items = grouped.get(severity, [])
                if not items:
                    continue
                    
                sev_color = severity_colors.get(severity, '#6B7280')
                sev_label = severity_labels.get(severity, severity)
                
                story.append(Paragraph(f"<font color='{sev_color}'><b>{sev_label} ({len(items)})</b></font>", styles['FindingTitle']))
                story.append(Spacer(1, 6))
                
                for idx, f in enumerate(items[:20], 1):  # Max 20 pro Kategorie
                    if not isinstance(f, dict):
                        continue
                    
                    rule_id = f.get('rule_id') or f.get('rule') or 'Unbekannt'
                    rule_name = rule_friendly_names.get(rule_id, rule_id)
                    message = f.get('message') or f.get('description') or ''
                    source_text = f.get('source') or f.get('src') or ''
                    target_text = f.get('target') or f.get('tgt') or ''
                    
                    # Befund-Titel
                    story.append(Paragraph(f"{idx}. {_rl_escape(rule_name)}", styles['FindingText']))
                    
                    # Nachricht
                    if message:
                        story.append(Paragraph(f"<i>{_rl_escape(str(message)[:200])}</i>", styles['FindingText']))
                    
                    # Quelltext / Übersetzung
                    if source_text or target_text:
                        if source_text:
                            src_display = str(source_text)[:150]
                            story.append(Paragraph(f"<b>Ausgangstext:</b> {_rl_escape(src_display)}", styles['SourceTarget']))
                        if target_text:
                            tgt_display = str(target_text)[:150]
                            story.append(Paragraph(f"<b>Übersetzung:</b> {_rl_escape(tgt_display)}", styles['SourceTarget']))
                    
                    # Zusätzliche Fehlerdetails
                    expected = f.get('expected')
                    found = f.get('found') or f.get('actual')
                    if expected or found:
                        if expected:
                            story.append(Paragraph(f"<font color='#10B981'>Erwartet: {_rl_escape(str(expected)[:100])}</font>", styles['ErrorText']))
                        if found:
                            story.append(Paragraph(f"<font color='#DC2626'>Gefunden: {_rl_escape(str(found)[:100])}</font>", styles['ErrorText']))
                    
                    story.append(Spacer(1, 8))
                
                if len(items) > 20:
                    story.append(Paragraph(f"<i>... und {len(items) - 20} weitere {sev_label}-Befunde</i>", styles['FindingText']))
                story.append(Spacer(1, 12))
        else:
            story.append(Paragraph("Befunde", styles['Section']))
            story.append(Paragraph("<font color='#10B981'>Keine Befunde gefunden – die Übersetzung scheint in Ordnung zu sein.</font>", styles['Normal']))
        
        # Empfehlungen
        recommendations = data.get('recommendations') if isinstance(data.get('recommendations'), list) else []
        if recommendations:
            story.append(Paragraph("Empfehlungen", styles['Section']))
            for rec in recommendations[:10]:
                if isinstance(rec, str):
                    story.append(Paragraph(f"• {_rl_escape(rec)}", styles['Normal']))
                elif isinstance(rec, dict):
                    rec_text = rec.get('text') or rec.get('message') or str(rec)
                    story.append(Paragraph(f"• {_rl_escape(str(rec_text))}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        # Fußzeile
        story.append(Spacer(1, 20))
        story.append(Paragraph("<i>Dieser Bericht wurde automatisch generiert.</i>", styles['Normal']))
        
        # PDF erstellen
        doc = SimpleDocTemplate(str(target), pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        doc.build(story)
        
        # Chart aufräumen
        if chart_path and chart_path.exists():
            try:
                chart_path.unlink(missing_ok=True)
            except Exception:
                pass
        
        return True
        
    except Exception as e:
        logger.error("PDF Export fehlgeschlagen: %s", e, exc_info=True)
        # Fallback: einfaches Canvas-PDF
        try:
            c = canvas.Canvas(str(target), pagesize=A4)
            width, height = A4
            y = height - 50
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, y, "Qualitätsanalyse-Bericht")
            y -= 30
            c.setFont("Helvetica", 10)
            if score is not None:
                c.drawString(50, y, f"Gesamtscore: {int(round(score))}/100")
                y -= 20
            c.drawString(50, y, f"Befunde: {len(findings)}")
            y -= 30
            for f in findings[:30]:
                if y < 60:
                    c.showPage()
                    y = height - 50
                if isinstance(f, dict):
                    text = f"- {f.get('rule_id', 'Befund')}: {str(f.get('message', ''))[:80]}"
                    c.drawString(60, y, text)
                    y -= 14
            c.save()
            return True
        except Exception:
            return False

# Rückwärtskompatibel
_export_pdf = export_pdf


def export_excel(report: Dict[str, Any] | Sequence[Dict[str, Any]], target: Path) -> bool:
    try:  # Lazy Import
        import openpyxl  # type: ignore
    except Exception:
        logger.warning("openpyxl nicht verfügbar – XLSX Export übersprungen")
        return False
    rows = _normalize(report)
    if not rows:
        logger.warning("Keine Daten für Excel Export")
        return False

    # Header: Reihenfolge aus erstem Row, dann neue Keys anhängen
    first_keys = list(rows[0].keys())
    seen = set(first_keys)
    for r in rows[1:]:
        for k in r.keys():
            if k not in seen:
                first_keys.append(k)
                seen.add(k)
    headers = first_keys

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    # Auto-Sizing Heuristik
    try:
        from openpyxl.utils import get_column_letter  # type: ignore
        for idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for r in rows:
                try:
                    v = r.get(h, "")
                    max_len = max(max_len, len(str(v)) if v is not None else 0)
                except Exception:
                    continue
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max_len + 2)
    except Exception:
        pass
    wb.save(target)
    return True

# Rückwärtskompatibel
_export_xlsx = export_excel


def _maybe_build_metrics_chart(report: Dict[str, Any] | Sequence[Dict[str, Any]]) -> Optional[Path]:
    """Erzeuge optionalen Chart (PNG) für Kernmetriken.

    Erwartete Keys: accuracy, fluency, style (0..1 oder 0..100). Nimmt erstes Element bei Sequenzen.
    """
    rows = _normalize(report)
    if not rows:
        return None
    metrics_row = rows[0]
    keys = ['accuracy', 'fluency', 'style']
    values: list[float] = []
    for k in keys:
        v = metrics_row.get(k)
        if v is None:
            return None  # Abbrechen wenn eine Kernmetrik fehlt -> kein Chart
        try:
            v_f = float(v)
            if v_f > 1.5:  # vermutlich Prozent (0..100)
                v_f = min(100.0, v_f) / 100.0
            values.append(max(0.0, min(1.0, v_f)))
        except Exception:
            return None
    try:
        import matplotlib
        matplotlib.use('Agg')  # Kopflose Backend erzwingen
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        logger.debug("matplotlib nicht verfügbar – kein Chart")
        return None
    try:
        import tempfile, uuid
        tmp_dir = Path(tempfile.gettempdir())
        chart_path = tmp_dir / f'quality_metrics_chart_{uuid.uuid4().hex}.png'
        fig, ax = plt.subplots(figsize=(4, 2.6))
        bars = ax.bar(keys, values, color=['#1F4E79', '#10B981', '#F59E0B'])  # vereinheitlichtes Brand-Blau
        ax.set_ylim(0, 1)
        ax.set_ylabel('Score')
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height()+0.02, f"{val*100:.1f}%", ha='center', va='bottom', fontsize=8)
        ax.set_title('Quality Metrics')
        plt.tight_layout()
        fig.savefig(chart_path, dpi=140)
        plt.close(fig)
        return chart_path
    except Exception:
        logger.debug("Chart Erstellung fehlgeschlagen", exc_info=True)
        return None

__all__ = [
    "export_analysis",
    "export_multi",
    "export_txt",
    "export_pdf",
    "export_excel",
    "SUPPORTED_FORMATS",
    "ExportResult",
]
